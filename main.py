"""
PROJECT: Duty Fairness Analyzer
AUTHOR: Rafael Cavalheiro
DESCRIPTION: 
    ETL Tool to parse Brazilian Military/Corporate PDF Rosters, normalize personnel names,
    and export statistical fairness analysis to Excel.
    UI: CustomTkinter (Dark Mode / Grayscale).
"""

import customtkinter as ctk
from tkinter import filedialog, messagebox
import os
import re
import sys
import threading
import pandas as pd
import pdfplumber
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image

# --- CONFIGURAÇÃO VISUAL ---
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("green")

# Função para encontrar recursos
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# --- CONSTANTES DE LÓGICA ---
TERMOS_A_REMOVER = [
    '1º TEN OTT', '2º TEN OTT', '1º OTT', '2º OTT', 'TEN CEL', 'TEN. CEL', 'MAJ MED', 'MAJ DENT',
    '1º SGT INF', '2º SGT INF', '3º SGT INF', '1º SGT COM', '3º SGT COM', '2º SGT MAT BEL',
    '3º SGT STT', '3º SGT SCT', 'CAP QCO', '1º TEN', '2º TEN', '1o TEN', '2o TEN', 'S TEN',
    '1º SGT', '2º SGT', '3º SGT', 'MAJ', 'CAP', 'TEN', 'SGT', 'ASP', 'CB', 'SD', 'OTT', 'QCO',
    'QEM', 'MED', 'INF', 'COM', 'STT', 'SCT', 'MAT BEL', 'º'
]

POSTOS_MAP = {
    "Sp Dia à Gu Campo Grande": "Sup de Dia à Gu Campo Grande",
    "Spvs Dia ao H Mil A CG": "Sup de Dia ao HMACG",
    "Veterinário de Sobreaviso à Gu CG": "Veterinário de Sobreaviso",
    "Cmt Gd ao Edifício Mello e Cáceres": "Cmt Gd Ed Mello e Cáceres",
    "Perito de Sobreaviso à Gu CG": "Perito de Sobreaviso",
    "Policial de Dia à Gu CG": "Policial de Dia",
    "Of Dia ao Forte Pantanal": "Of Dia ao Forte Pantanal",
    "Of Dia ao Forte Pantanal (Aprendiz):": "Of Dia ao Forte Pantanal (Aprendiz)",
    "Adj Of Dia": "Adj Of Dia",
    "Cmt Gd ao Forte Pantanal": "Cmt Gd ao Forte Pantanal",
    "Aux Cmt Gd ao Forte Pantanal": "Aux Cmt Gd ao Forte Pantanal",
    "Cb Gd ao Forte Pantanal": "Cb Gd ao Forte Pantanal",
    "Aux Cb Gd ao Forte Pantanal": "Aux Cb Gd ao Forte Pantanal",
    "Corneteiro": "Corneteiro"
}

class DutyFairnessApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        
        # Configuração da Janela
        self.title("Duty Fairness Analyzer")
        self.geometry("600x480")
        self.resizable(False, False)
        
        # Regex Compilado
        place_pattern = "|".join(re.escape(place) for place in POSTOS_MAP.keys())
        self.place_regex = re.compile(rf"({place_pattern}):\s*(.*?)\s*-\s*([A-Z0-9º\s/.]+)")
        self.date_regex = re.compile(r"PARA O DIA\s+(\d{1,2})\s+DE\s+([A-ZÇ]+)\s+DE\s+(\d{4})\s+\(([A-ZÀ-Ú-]+)\)")
        
        # Ícone da Janela
        try:
            self.iconbitmap(resource_path("jacare.ico"))
        except: pass

        self._setup_ui()

    def _setup_ui(self):
        # --- HEADER ---
        self.lbl_title = ctk.CTkLabel(self, text="Duty Fairness Analyzer", font=("Roboto", 24, "bold"))
        self.lbl_title.pack(pady=(25, 5))
        
        self.lbl_subtitle = ctk.CTkLabel(self, text="Auditoria de Escalas e Equidade (PDF -> Excel)", font=("Roboto", 12), text_color="gray")
        self.lbl_subtitle.pack(pady=(0, 20))

        # --- ÁREA DE SELEÇÃO ---
        self.frame_main = ctk.CTkFrame(self, fg_color="#181818") # Fundo mais escuro
        self.frame_main.pack(pady=10, padx=20, fill="x")

        # Input PDF
        self.pdf_path = ctk.StringVar()
        
        # Carregar ícones (opcional, fallback para texto)
        try:
            img_folder = ctk.CTkImage(Image.open(resource_path("folder_ico.png")), size=(20, 20))
            img_excel = ctk.CTkImage(Image.open(resource_path("excel_icon.png")), size=(20, 20))
        except:
            img_folder, img_excel = None, None

        self.btn_pdf = ctk.CTkButton(
            self.frame_main, 
            text="Selecionar Pasta de Boletins (PDF)", 
            image=img_folder, 
            compound="left",
            command=self.selecionar_pasta_pdf,
            fg_color="#2B2B2B", 
            hover_color="#3A3A3A", 
            height=40,
            font=("Roboto", 13)
        )
        self.btn_pdf.pack(fill="x", pady=(15, 5), padx=15)

        self.lbl_pdf_path = ctk.CTkLabel(self.frame_main, text="Nenhuma pasta selecionada", font=("Roboto", 11), text_color="gray")
        self.lbl_pdf_path.pack(pady=(0, 15), padx=20, anchor="w")

        # Input Excel
        self.xlsx_path = ctk.StringVar()
        self.btn_xlsx = ctk.CTkButton(
            self.frame_main, 
            text="Selecionar Destino do Relatório (Excel)", 
            image=img_excel, 
            compound="left",
            command=self.selecionar_pasta_xlsx,
            fg_color="#2B2B2B", 
            hover_color="#3A3A3A", 
            height=40,
            font=("Roboto", 13)
        )
        self.btn_xlsx.pack(fill="x", pady=(5, 5), padx=15)

        self.lbl_xlsx_path = ctk.CTkLabel(self.frame_main, text="Nenhuma pasta selecionada", font=("Roboto", 11), text_color="gray")
        self.lbl_xlsx_path.pack(pady=(0, 15), padx=20, anchor="w")

        # --- CONTROLES DE AÇÃO ---
        
        # Barra de Progresso
        self.progress = ctk.CTkProgressBar(self, height=10, width=400)
        self.progress.set(0) # 0%
        self.progress.pack(pady=(20, 10))

        # Botão Principal
        self.btn_action = ctk.CTkButton(
            self, 
            text="PROCESSAR DADOS", 
            font=("Roboto", 14, "bold"), 
            height=50,
            fg_color="#27ae60",
            hover_color="#2ecc71",
            command=self.start_thread
        )
        self.btn_action.pack(pady=10, padx=40, fill="x")

    # --- LÓGICA DE UI ---
    def selecionar_pasta_pdf(self):
        pasta = filedialog.askdirectory(title="Selecione a pasta com os Boletins (PDF)")
        if pasta:
            self.pdf_path.set(pasta)
            self.lbl_pdf_path.configure(text=f".../{os.path.basename(pasta)}", text_color="#E0E0E0")

    def selecionar_pasta_xlsx(self):
        pasta = filedialog.askdirectory(title="Selecione onde salvar o Excel")
        if pasta:
            self.xlsx_path.set(pasta)
            self.lbl_xlsx_path.configure(text=f".../{os.path.basename(pasta)}", text_color="#E0E0E0")

    # --- THREADING ---
    def start_thread(self):
        folder_path = self.pdf_path.get()
        output_dir = self.xlsx_path.get()

        if not folder_path or not output_dir:
            messagebox.showwarning("Atenção", "Selecione as pastas de origem e destino.")
            return

        self.btn_action.configure(state="disabled", text="PROCESSANDO...", fg_color="#555555")
        self.progress.configure(mode="indeterminate")
        self.progress.start()
        
        threading.Thread(target=self.run_logic, args=(folder_path, output_dir), daemon=True).start()

    def run_logic(self, folder_path, output_dir):
        try:
            output_file = self.processar_arquivos(folder_path, output_dir)
            self.after(0, lambda: self.finish_success(output_file))
        except Exception as e:
            self.after(0, lambda: self.finish_error(str(e)))

    def finish_success(self, output_file):
        self.progress.stop()
        self.progress.configure(mode="determinate")
        self.progress.set(1) # 100%
        self.btn_action.configure(state="normal", text="PROCESSAR DADOS", fg_color="#27ae60")
        messagebox.showinfo("Sucesso", f"Relatório gerado:\n{output_file}")

    def finish_error(self, error_msg):
        self.progress.stop()
        self.progress.set(0)
        self.btn_action.configure(state="normal", text="PROCESSAR DADOS", fg_color="#27ae60")
        messagebox.showerror("Erro Crítico", f"Falha ao processar:\n{error_msg}")

    # --- LÓGICA ETL---
    def _padronizar_nome(self, nome_bruto):
        nome_limpo = nome_bruto.upper()
        for termo in TERMOS_A_REMOVER:
            nome_limpo = nome_limpo.replace(termo, ' ')
        return re.sub(r'\s+', ' ', nome_limpo).strip() or nome_bruto.strip()

    def processar_arquivos(self, folder_path, output_dir):
        output_file = os.path.join(output_dir, "Relatorio_Escala_Justa.xlsx")
        meses = {'JANEIRO': '01', 'FEVEREIRO': '02', 'MARÇO': '03', 'ABRIL': '04', 'MAIO': '05', 'JUNHO': '06', 'JULHO': '07', 'AGOSTO': '08', 'SETEMBRO': '09', 'OUTUBRO': '10', 'NOVEMBRO': '11', 'DEZEMBRO': '12'}

        all_data = []
        pdf_files = [f for f in os.listdir(folder_path) if f.lower().endswith(".pdf")]
        
        if not pdf_files: raise ValueError("Nenhum arquivo PDF encontrado.")

        for filename in pdf_files:
            file_path = os.path.join(folder_path, filename)
            with pdfplumber.open(file_path) as pdf:
                full_text = "".join(page.extract_text(x_tolerance=2) or "" for page in pdf.pages)
                
                date_matches = list(self.date_regex.finditer(full_text))
                if not date_matches:
                    all_data.append({"Arquivo": filename, "Data": "N/A", "Dia": "N/A", "Posto": "N/A", "Nome": "N/A", "Companhia": "N/A"})
                    continue

                for i, date_match in enumerate(date_matches):
                    day_num, month_name, year, day_of_week = date_match.groups()
                    date = f"{day_num.zfill(2)}/{meses.get(month_name, 'XX')}/{year}"
                    
                    start_pos = date_match.end()
                    end_pos = date_matches[i + 1].start() if i + 1 < len(date_matches) else len(full_text)
                    section_text = full_text[start_pos:end_pos]
                    
                    for line in section_text.split('\n'):
                        place_match = self.place_regex.search(line)
                        if place_match:
                            posto_bruto, nome_bruto, companhia = place_match.groups()
                            posto_padronizado = POSTOS_MAP.get(posto_bruto.strip(), posto_bruto.strip())
                            all_data.append({
                                "Arquivo": filename,
                                "Data": date,
                                "Dia": day_of_week,
                                "Posto": posto_padronizado,
                                "Nome": nome_bruto.strip(),
                                "Companhia": companhia.strip().strip('.')
                            })

        if not all_data: raise ValueError("Não foi possível extrair dados válidos.")

        df = pd.DataFrame(all_data)
        df_analise = df[(df['Nome'] != 'N/A') & (df['Nome'] != 'Unknown')].copy()
        
        if not df_analise.empty:
            df_analise['Nome_Chave'] = df_analise['Nome'].apply(self._padronizar_nome)
            total_dias_por_posto = df_analise.groupby('Posto')['Data'].nunique()
            total_dias_geral = df['Data'][df['Data'] != 'Data não encontrada'].nunique()

            def aggregate_posts(series):
                post_counts = series.value_counts()
                details = []
                for post, count in post_counts.items():
                    total_days = total_dias_por_posto.get(post, 1)
                    percentage = (count / total_days) * 100 if total_days > 0 else 0
                    details.append(f"{post}: {count}/{total_days} ({percentage:.1f}%)")
                return " | ".join(sorted(details))

            grouped = df_analise.groupby('Nome_Chave').agg(
                Nome_Exibicao=('Nome', 'first'),
                Contagem=('Nome', 'count'),
                Postos_Detalhes=('Posto', aggregate_posts)
            ).reset_index()
            
            grouped = grouped.rename(columns={'Nome_Exibicao': 'Nome', 'Postos_Detalhes': 'Detalhes'})
            del grouped['Nome_Chave']
            grouped = grouped.sort_values('Contagem', ascending=False)
            
            summary_df = pd.DataFrame({'Nome': [f'Total Dias: {total_dias_geral}'], 'Contagem': [''], 'Detalhes': ['']})
            final_grouped = pd.concat([summary_df, grouped], ignore_index=True)
        else:
            df = pd.DataFrame(columns=["Arquivo", "Data", "Dia", "Posto", "Nome", "Companhia"])
            final_grouped = pd.DataFrame()

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Dados Brutos')
            final_grouped.iloc[1:].to_excel(writer, index=False, sheet_name='Analise', startrow=1)
            self.format_sheet(writer.sheets['Dados Brutos'], df)
            self.format_sheet(writer.sheets['Analise'], final_grouped, is_summary_sheet=True)
            
        return output_file

    def format_sheet(self, worksheet, dataframe, is_summary_sheet=False):
        # Formatação Excel
        header_font = Font(bold=True, color="FFFFFF", name='Calibri')
        header_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
        thin_border = Border(left=Side(style='thin', color="DCDCDC"), right=Side(style='thin', color="DCDCDC"), top=Side(style='thin', color="DCDCDC"), bottom=Side(style='thin', color="DCDCDC"))
        
        start_row = 1
        if is_summary_sheet and not dataframe.empty:
            start_row = 2
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(dataframe.columns))
            worksheet['A1'].value = dataframe.iloc[0, 0]
            worksheet['A1'].fill = PatternFill(start_color="D3D3D3", fill_type="solid")
            worksheet['A1'].font = Font(bold=True)
            worksheet['A1'].alignment = Alignment(horizontal='center')

        for col_idx, column in enumerate(dataframe.columns, 1):
            cell = worksheet.cell(row=start_row, column=col_idx)
            cell.value = column
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            max_len = max([len(str(val)) for val in dataframe[column].values] + [len(column)])
            worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

if __name__ == "__main__":
    app = DutyFairnessApp()
    app.mainloop()
